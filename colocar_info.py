'''
FORMATAÇÃO DA TABELA:

A = Plaqueta    (int)
B = Descrição   (str)
C = Marca       (str)
D = Modelo      (str)
E = Cor         (str)
F = Unidade     (int)
G = Observação  (str)
H = local       (str)


Até a linha 901 Unidade é '1' (sede);
De 902 até a linha 1104 Unidade é '13' (forum civel ou criminal)
De 1105 até a linha 1111 Unidade é '14' (cityhall)
De 1112 até a linha 1114 Unidade é '15' (VEP)

doc1 = {
    "unidade" : 1,
    "local" : 'sede',
    "promotor": 'Ricardo',
    "atribuicao": 'coordenação administrativa'}
'''

from contextlib import nullcontext
from pymongo import MongoClient
from datetime import date

CONNECTION_STRING = "mongodb://localhost:27017"
client = MongoClient(CONNECTION_STRING)
SGP_db = client["SGP_db"]
unidades = SGP_db["Unidades"]
promotoria = SGP_db["Unidades{'promotoria': 1}"]

from distutils.command.build import build
from openpyxl import load_workbook
 
ROWS, COLS = 38, 5
WORKBOOK, WORKSHEET = "promotorias.xlsx", "Sheet1"
#WORKBOOK, WORKSHEET = "LONDRINA - ADM.xlsx", "Edição para Banco de dados"
wb = load_workbook(WORKBOOK)
ws = wb[WORKSHEET]

todosPatrimonios = {}
class inserirDadosSLSX ():
    def __init__(self, linha):

        self.cunidade = ws['F'+ str(linha)]
        self.cplaqueta = ws['A'+ str(linha)]
        self.cdescicao = ws['B'+ str(linha)]
        self.cmarca = ws['C'+ str(linha)]
        self.cmodelo = ws['D'+ str(linha)]
        self.ccor = ws['E'+ str(linha)]
        self.cobservacao = ws['G'+ str(linha)]

        self.unidade = ws['A' + str(linha)]
        self.nome = ws['B' + str(linha)]
        self.local = ws['C' + str(linha)]
        self.atribuicao = ws['D' + str(linha)]
        self.promotor = ws['E' + str(linha)]

    def inserirunidade(self):

        self.dictunidade  = {
            'unidade': self.unidade.value,
            'nome': self.nome.value,
            'local' : self.local.value
        }

        if self.atribuicao != None:
            self.dictunidade['atribuicao'] = self.atribuicao.value
        
        if self.promotor != None:
            self.dictunidade['promotor(a)'] = self.promotor.value
            
        unidades.insert_one(
            self.dictunidade
        )

    def inserirPatrimonio(self):

        self.dictPatrimonio  = {
            'plaqueta': self.cplaqueta.value,
            'descricao': self.cdescicao.value
        }

        if self.cmarca.value != None:
            self.dictPatrimonio['marca'] = self.cmarca.value

        if self.cmodelo.value != None:
            self.dictPatrimonio['modelo'] = self.cmodelo.value
    
        if self.ccor.value != None:
            self.dictPatrimonio['cor'] = self.ccor.value

        if self.cobservacao.value != None:
            self.dictPatrimonio['observacao'] = self.cobservacao.value

        unidades.update_one(
            {"unidade" : self.cunidade.value,},
            {'$set':{str(self.cplaqueta.value): self.dictPatrimonio}}
        )

i = 2
while i <= 38:
    print('inserirndo unidade: ', i)
    unidade = inserirDadosSLSX(i)
    unidade.inserirunidade()
    i += 1
    del unidade

wb.close()

ROWS, COLS = 1114, 8
#WORKBOOK, WORKSHEET = "promotorias.xlsx", "Sheet1"
WORKBOOK, WORKSHEET = "LONDRINA - ADM.xlsx", "Edição para Banco de dados"
wb = load_workbook(WORKBOOK)
ws = wb[WORKSHEET]

aux = 2
i = aux

while i <= 901:
    print('inserirndo linha: ', i)
    patrimonio = inserirDadosSLSX(i)
    patrimonio.inserirPatrimonio()
    i = i+1
    del patrimonio

'''
'''